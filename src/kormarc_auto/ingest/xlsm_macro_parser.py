"""Excel xlsm 매크로 파일에서 ISBN 추출.

Part 50 발견: 영업 자료 27건 (P1 매크로 사서) ↔ 코드 불일치 해결.
ADR-0070.

검증된 라이브러리: openpyxl (xlsm 지원, 매크로는 무시·셀 데이터만 추출).
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from pathlib import Path

logger = logging.getLogger(__name__)

# ISBN-13 패턴 (978·979 시작 13자리, 하이픈 허용)
ISBN_PATTERN = re.compile(r"\b97[89][\d-]{10,17}\b")


@dataclass
class XlsmExtractionResult:
    """xlsm 추출 결과."""

    file_path: Path
    total_rows: int
    isbn_rows: int
    isbns: list[str]
    detected_isbn_columns: list[str]
    warnings: list[str]


def extract_isbns_from_xlsm(
    file_path: str | Path, *, max_isbns: int = 1000
) -> XlsmExtractionResult:
    """xlsm 파일에서 ISBN 자동 추출 (매크로 무시·데이터만).

    P1 매크로 사서 워크플로:
    1. 자작 매크로 .xlsm 그대로 업로드
    2. 모든 시트·컬럼에서 ISBN-13 패턴 자동 검출
    3. 휴리스틱으로 ISBN 컬럼 추론 (제목·헤더에 "ISBN", "도서번호" 등)
    4. 추출 결과 사서에게 미리보기 → 확인 후 일괄 처리

    Args:
        file_path: xlsm·xlsx 파일 경로
        max_isbns: 안전 상한 (기본 1,000건)

    Returns:
        XlsmExtractionResult — ISBN 목록 + 감지 컬럼 + 경고
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return XlsmExtractionResult(
            file_path=Path(file_path),
            total_rows=0,
            isbn_rows=0,
            isbns=[],
            detected_isbn_columns=[],
            warnings=["openpyxl 미설치 — pip install openpyxl 필요"],
        )

    path = Path(file_path)
    if not path.exists():
        return XlsmExtractionResult(
            file_path=path,
            total_rows=0,
            isbn_rows=0,
            isbns=[],
            detected_isbn_columns=[],
            warnings=[f"파일을 찾을 수 없어요: {path}"],
        )

    warnings: list[str] = []
    isbns: list[str] = []
    detected_columns: set[str] = set()
    total_rows = 0

    try:
        wb = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    except Exception as e:
        return XlsmExtractionResult(
            file_path=path,
            total_rows=0,
            isbn_rows=0,
            isbns=[],
            detected_isbn_columns=[],
            warnings=[f"xlsm 파일 읽기 실패: {e}"],
        )

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws is None:
            continue

        # 헤더 행 분석 (1행 또는 2행)
        header_isbn_cols: list[int] = []
        for header_row in (1, 2):
            try:
                row_cells = next(
                    ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
                )
            except StopIteration:
                continue

            for col_idx, cell in enumerate(row_cells):
                if cell is None:
                    continue
                cell_str = str(cell).lower()
                if any(
                    keyword in cell_str for keyword in ("isbn", "도서번호", "도서 번호", "도서코드")
                ):
                    header_isbn_cols.append(col_idx)
                    detected_columns.add(f"{sheet_name}!{cell}")

        # 데이터 행 ISBN 추출
        for row in ws.iter_rows(min_row=2, values_only=True):
            total_rows += 1
            if total_rows > max_isbns * 10:  # 무한 루프 방지
                warnings.append(f"행 한도 초과 ({max_isbns * 10}행) — 일부만 추출")
                break

            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                cell_str = str(cell)

                # 헤더로 검출된 ISBN 컬럼 우선 처리
                if col_idx in header_isbn_cols:
                    isbn = _normalize_isbn(cell_str)
                    if isbn:
                        isbns.append(isbn)
                        if len(isbns) >= max_isbns:
                            break
                else:
                    # 모든 셀에서 ISBN 패턴 휴리스틱 검출
                    matches = ISBN_PATTERN.findall(cell_str)
                    for match in matches:
                        isbn = _normalize_isbn(match)
                        if isbn:
                            isbns.append(isbn)
                            if len(isbns) >= max_isbns:
                                break

            if len(isbns) >= max_isbns:
                break

    wb.close()

    if not detected_columns:
        warnings.append("ISBN 헤더 컬럼을 자동 감지하지 못했어요. 셀 패턴 검색으로 추출했습니다.")

    if not isbns:
        warnings.append(
            "ISBN을 찾지 못했어요. ISBN-13 (978·979 시작 13자리)이 셀에 있는지 확인해주세요."
        )

    # 중복 제거 (순서 유지)
    seen = set()
    unique_isbns: list[str] = []
    for isbn in isbns:
        if isbn not in seen:
            seen.add(isbn)
            unique_isbns.append(isbn)

    return XlsmExtractionResult(
        file_path=path,
        total_rows=total_rows,
        isbn_rows=len(unique_isbns),
        isbns=unique_isbns,
        detected_isbn_columns=sorted(detected_columns),
        warnings=warnings,
    )


def _normalize_isbn(raw: str) -> str | None:
    """ISBN-13 정규화 (하이픈 제거 + 체크섬 검증).

    Returns:
        13자리 정규화 ISBN 또는 None (유효하지 않음)
    """
    digits = re.sub(r"[^\d]", "", raw)
    if len(digits) != 13:
        return None
    if not digits.startswith(("978", "979")):
        return None

    # ISBN-13 체크섬 검증
    total = sum(int(d) * (1 if i % 2 == 0 else 3) for i, d in enumerate(digits[:12]))
    check = (10 - total % 10) % 10
    if check != int(digits[12]):
        return None

    return digits
