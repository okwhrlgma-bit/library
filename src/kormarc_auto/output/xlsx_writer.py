"""Excel(.xlsx) 출력 — 사서가 가장 많이 쓰는 자료 형식.

KOLAS·DLS·KOLISYS-NET 외에 사서들이 일상적으로 쓰는 양식.
.mrc는 KOLAS 반입용, .xlsx는 사서 자체 관리·검토·납본 명세·일괄 편집용.

출력 시트:
1. 메인: 1행 1책 — 020/100/245/260/300/490/500/650/856 등 흔한 필드
2. 원본: pymarc Record를 mrk(텍스트) 형식 그대로 (고급 사서 검토)
3. 메타: confidence·source·생성일시
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


# 사서가 평소에 자주 쓰는 BookData 키 → Excel 컬럼명
DEFAULT_COLUMNS = [
    ("isbn", "ISBN"),
    ("title", "표제"),
    ("subtitle", "부서명"),
    ("author", "저자"),
    ("publisher", "발행자"),
    ("publication_place", "발행지"),
    ("publication_year", "발행연도"),
    ("edition", "판차"),
    ("series", "총서"),
    ("pages", "페이지"),
    ("size", "크기"),
    ("price", "가격"),
    ("kdc", "KDC"),
    ("ddc", "DDC"),
    ("call_number", "청구기호"),
    ("registration_no", "등록번호"),
    ("subject", "주제어"),
    ("summary", "내용요약"),
    ("url", "URL"),
    ("confidence", "신뢰도"),
    ("source", "데이터 출처"),
]


def write_books_xlsx(
    books: list[dict[str, Any]],
    *,
    output_path: str | Path,
    sheet_name: str = "도서목록",
    columns: list[tuple[str, str]] | None = None,
) -> Path:
    """BookData dict 리스트 → 1행 1책 xlsx.

    사서가 KOLAS 반입 전 검토·일괄 편집·납본 명세 작성에 사용.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:
        raise RuntimeError("openpyxl 미설치 — `pip install openpyxl`") from e

    cols = columns or DEFAULT_COLUMNS
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 헤더 (한국어 컬럼명)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    for col_idx, (_, label) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 본문
    for row_idx, book in enumerate(books, 2):
        for col_idx, (key, _) in enumerate(cols, 1):
            value = book.get(key, "")
            if isinstance(value, list):
                value = "; ".join(str(v) for v in value)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 컬럼 폭 자동
    for col_idx, (_, label) in enumerate(cols, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(
            10, min(40, len(label) * 2 + 5)
        )

    # 메타데이터 시트
    meta = wb.create_sheet("메타")
    meta["A1"] = "생성"
    meta["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    meta["A2"] = "총 건수"
    meta["B2"] = len(books)
    meta["A3"] = "도구"
    meta["B3"] = "kormarc-auto"
    meta["A4"] = "주의"
    meta["B4"] = "사서 검토 필수 — AI 추천 후보는 정확도 보장하지 않음"

    wb.save(out)
    logger.info("xlsx 저장: %s (%d건)", out, len(books))
    return out


def write_isbn_template_xlsx(*, output_path: str | Path) -> Path:
    """사서가 ISBN을 채워 넣을 수 있는 빈 템플릿.

    A열: ISBN (사서가 채움)
    B~Z열: 자동 채워질 자리 (Power Query·VBA 매크로 또는 batch CLI 후처리)
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:
        raise RuntimeError("openpyxl 미설치 — `pip install openpyxl`") from e

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ISBN 입력"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")

    headers = ["ISBN (필수)", *[label for _, label in DEFAULT_COLUMNS[1:]]]
    for col_idx, label in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 안내 시트
    guide = wb.create_sheet("사용법")
    guide["A1"] = "kormarc-auto Excel 템플릿 사용법"
    guide["A1"].font = Font(bold=True, size=14)
    guide["A3"] = "1. 'ISBN 입력' 시트 A열에 ISBN을 입력 (한 줄에 하나)"
    guide["A4"] = "2. 저장 후 다음 명령으로 일괄 처리:"
    guide["A5"] = "   kormarc-auto batch-xlsx input.xlsx --output filled.xlsx"
    guide["A6"] = "3. 또는 Power Query로 REST API 호출 (docs/excel-integration.md 참조)"
    guide["A8"] = "주의: AI 추천 결과는 사서 검토 필수입니다."

    for col_idx, _ in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

    wb.save(out)
    logger.info("xlsx 템플릿 저장: %s", out)
    return out


__all__ = ["DEFAULT_COLUMNS", "write_books_xlsx", "write_isbn_template_xlsx"]
