"""제적·폐기 결재서식 PDF + 폐기목록 엑셀 — 장서개발지침(제3판) §3.2.

근거:
- 도서관법 §22 (자료 제적·폐기) + 동 시행령 §16
- 장서개발지침 제3장 §3.2 "자료제적·폐기 세부지침" — 분기·연간 제적심의

사서 워크플로우:
1. 제적 후보 자료 식별 (노후·중복·이용저조·분실·파손)
2. 583 처리 코드 부여 (PRESERVE/WITHDRAW 등)
3. 제적심의 결재서식 작성 → 운영위 의결 → 폐기목록 행정 보존
4. 등록번호 회수 (registration.find_missing_numbers와 협력)

기존 `kormarc/loss_damage.py`는 583 코드 부여까지만. 본 모듈은 결재서식·목록 PDF/XLSX.
"""

from __future__ import annotations

import logging
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)


DISPOSAL_REASONS = {
    "WORN": "노후·파손 (이용 불가)",
    "DUPL": "복본 과다 (3부 초과)",
    "OBSO": "내용 진부화 (자료 가치 상실)",
    "LOWUSE": "장기 미이용 (5년 이상 대출 0회)",
    "LOST": "분실 (장서점검 미발견)",
    "DAMAGED": "심각 파손 (수리 불가)",
    "REPLACE": "신판 교체 (구판 폐기)",
    "OTHER": "기타 (운영위 결의)",
}


@dataclass
class DisposalEntry:
    """제적 1건."""

    registration_number: str
    title: str
    author: str | None = None
    publisher: str | None = None
    publication_year: str | None = None
    isbn: str | None = None
    kdc: str | None = None
    call_number: str | None = None
    reason_code: str = "OTHER"  # DISPOSAL_REASONS의 키
    reason_detail: str = ""
    inspector: str = ""  # 점검 사서명
    cost_krw: int | None = None  # 매입가 (보고용)


def render_disposal_form_pdf(
    entries: Iterable[DisposalEntry],
    *,
    library_name: str,
    fiscal_period: str,  # "2026 1분기" 등
    director: str = "",
    committee: str = "도서관운영위원회",
    output_path: str | Path | None = None,
) -> Path:
    """제적심의 결재서식 PDF.

    형식: 결재선(작성자·검토·결재자) + 제적 요약 + 자료 목록 + 운영위 의결.
    인쇄 후 사서·관장·운영위원이 서명 → 행정 보존.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
    except ImportError as e:
        raise RuntimeError("reportlab 미설치 — `pip install reportlab`") from e

    from kormarc_auto.output.reports import _korean_font

    entries_list = list(entries)
    out = Path(
        output_path
        or f"logs/disposal/disposal_{fiscal_period.replace(' ', '_')}_{datetime.now():%Y%m%d}.pdf"
    )
    out.parent.mkdir(parents=True, exist_ok=True)

    c = canvas.Canvas(str(out), pagesize=A4)
    font = _korean_font(c)
    page_w, page_h = A4
    margin = 20 * mm
    y = page_h - margin

    # 표제
    c.setFont(font, 18)
    c.drawString(margin, y, "도서 제적·폐기 결재서식")
    y -= 8 * mm
    c.setFont(font, 10)
    c.drawString(margin, y, "(도서관법 §22 + 장서개발지침 §3.2)")
    y -= 10 * mm

    # 메타
    c.setFont(font, 12)
    c.drawString(margin, y, f"도서관: {library_name}")
    y -= 6 * mm
    c.drawString(margin, y, f"심의 기간: {fiscal_period}")
    y -= 6 * mm
    c.drawString(margin, y, f"제적 건수: {len(entries_list)}건")
    y -= 6 * mm
    total_cost = sum(e.cost_krw or 0 for e in entries_list)
    if total_cost:
        c.drawString(margin, y, f"매입가 합계(참고): {total_cost:,}원")
        y -= 6 * mm
    y -= 6 * mm

    # 결재선
    c.setFont(font, 11)
    c.drawString(margin, y, "[결재선]")
    y -= 7 * mm
    boxes = [
        ("작성", entries_list[0].inspector if entries_list else ""),
        ("검토", "사서장"),
        ("결재", director or "관장"),
        ("의결", committee),
    ]
    box_w = (page_w - 2 * margin) / 4 - 2 * mm
    for i, (label, name) in enumerate(boxes):
        x = margin + i * (box_w + 2 * mm)
        c.rect(x, y - 22 * mm, box_w, 22 * mm)
        c.setFont(font, 10)
        c.drawString(x + 2, y - 5 * mm, label)
        c.setFont(font, 9)
        c.drawString(x + 2, y - 10 * mm, name)
        c.drawString(x + 2, y - 18 * mm, "(인)")
    y -= 28 * mm

    # 사유 요약
    c.setFont(font, 12)
    c.drawString(margin, y, "[제적 사유 요약]")
    y -= 7 * mm
    c.setFont(font, 10)
    reason_count: dict[str, int] = {}
    for e in entries_list:
        reason_count[e.reason_code] = reason_count.get(e.reason_code, 0) + 1
    for code, cnt in sorted(reason_count.items(), key=lambda x: -x[1]):
        label = DISPOSAL_REASONS.get(code, code)
        c.drawString(margin + 5 * mm, y, f"· {code} {label}: {cnt}건")
        y -= 5 * mm
    y -= 6 * mm

    # 목록 헤더
    c.setFont(font, 11)
    c.drawString(margin, y, "[제적 자료 목록]")
    y -= 7 * mm
    headers = ["등록번호", "표제", "사유"]
    col_x = [margin, margin + 35 * mm, margin + 130 * mm]
    c.setFont(font, 9)
    for h, x in zip(headers, col_x, strict=False):
        c.drawString(x, y, h)
    y -= 5 * mm
    c.line(margin, y, page_w - margin, y)
    y -= 3 * mm

    # 자료 목록 (페이지 넘김)
    for e in entries_list:
        if y < margin + 20 * mm:
            c.showPage()
            y = page_h - margin
            c.setFont(font, 9)
        c.drawString(col_x[0], y, e.registration_number[:14])
        title = (e.title or "")[:40]
        c.drawString(col_x[1], y, title)
        c.drawString(col_x[2], y, e.reason_code)
        y -= 5 * mm

    # 운영위 의결문
    if y < margin + 30 * mm:
        c.showPage()
        y = page_h - margin
    y -= 8 * mm
    c.setFont(font, 11)
    c.drawString(margin, y, "[운영위원회 의결]")
    y -= 7 * mm
    c.setFont(font, 10)
    c.drawString(
        margin,
        y,
        f"위 {len(entries_list)}건의 자료를 도서관법 §22 및 본 도서관 장서개발지침에 따라",
    )
    y -= 5 * mm
    c.drawString(margin, y, "제적·폐기함을 운영위원회에서 의결함.")
    y -= 12 * mm
    c.drawString(margin, y, f"의결일: {datetime.now():%Y년 %m월 %d일}")

    # 푸터
    c.setFont(font, 8)
    c.drawString(margin, margin, f"발행: kormarc-auto · {datetime.now():%Y-%m-%d %H:%M}")
    c.save()
    logger.info("제적·폐기 PDF 저장: %s (%d건)", out, len(entries_list))
    return out


def write_disposal_xlsx(
    entries: Iterable[DisposalEntry],
    output_path: str | Path,
) -> Path:
    """폐기목록 엑셀 — 행정 보존용. 14컬럼."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as e:
        raise RuntimeError("openpyxl 미설치 — `pip install openpyxl`") from e

    cols = [
        "등록번호",
        "표제",
        "저자",
        "출판사",
        "발행연도",
        "ISBN",
        "KDC",
        "청구기호",
        "사유코드",
        "사유설명",
        "사유상세",
        "점검자",
        "매입가(원)",
        "비고",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "제적·폐기 목록"
    ws.append(cols)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for e in entries:
        ws.append(
            [
                e.registration_number,
                e.title,
                e.author or "",
                e.publisher or "",
                e.publication_year or "",
                e.isbn or "",
                e.kdc or "",
                e.call_number or "",
                e.reason_code,
                DISPOSAL_REASONS.get(e.reason_code, ""),
                e.reason_detail,
                e.inspector,
                e.cost_krw or "",
                "",
            ]
        )

    from openpyxl.utils import get_column_letter

    for i, col_name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(col_name) + 4)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    logger.info("폐기목록 XLSX: %s", out)
    return out


__all__ = [
    "DISPOSAL_REASONS",
    "DisposalEntry",
    "render_disposal_form_pdf",
    "write_disposal_xlsx",
]
