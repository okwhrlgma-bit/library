"""연간 KOLIS-NET 통계 — 도서관법 §22의2(국가서지·통계) + 도서관운영평가.

근거:
- 도서관법 §22의2 (국가서지·통계 협조)
- 도서관운영평가 평가지표 (한국도서관협회·문화체육관광부)
- 학교도서관진흥법 §6 (한국교육학술정보원 정보서비스시스템 연계)
- 「국가도서관통계시스템 KOLIS-NET」 양식

연간 통계 4영역:
1. 장서 구성 (자료유형·KDC 분류·언어)
2. 수서 (구입·기증·납본 권수·금액)
3. 이용 (대출·이용자·방문)  ← 우리 자관 인덱스 + 외부 입력
4. 운영 (휴관일·개관시간·인력)

PO 자관 인덱스(`inventory.library_db`)에서 1·2영역 자동 산출.
3·4는 사서 입력(엑셀 양식 빈 칸).
"""

from __future__ import annotations

import logging
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


# 자료유형 코드 (KOLIS-NET 표준)
MATERIAL_TYPES = [
    ("BK", "도서"),
    ("SE", "연속간행물"),
    ("MM", "비도서(멀티미디어)"),
    ("EB", "전자책"),
    ("EJ", "전자저널"),
    ("RB", "고서"),
    ("ETC", "기타"),
]


@dataclass
class AnnualStats:
    """연간 통계 1년치."""

    year: int
    library_name: str
    library_code: str = ""  # KOLIS-NET 도서관 코드
    holdings_total: int = 0
    holdings_by_kdc: dict[str, int] = field(default_factory=dict)
    holdings_by_material: dict[str, int] = field(default_factory=dict)
    acquisition_purchase: int = 0  # 구입 권수
    acquisition_donation: int = 0  # 기증
    acquisition_deposit: int = 0  # 납본
    acquisition_cost_krw: int = 0
    loans_total: int = 0  # 사서 입력
    visits_total: int = 0  # 사서 입력
    members_total: int = 0  # 사서 입력
    operating_days: int = 0
    closed_days: int = 0
    librarians: int = 0  # 정규직 사서
    notes: str = ""


def compute_holdings_from_index(
    *,
    year: int | None = None,
) -> tuple[int, dict[str, int], dict[str, int]]:
    """자관 DB 인덱스에서 장서 통계 산출.

    Returns:
        (총 장서 수, KDC 1자리 분포, 자료유형 분포)
    """
    try:
        from kormarc_auto.inventory.library_db import search_local
    except ImportError:
        return 0, {}, {}

    records = search_local("", limit=1000000)
    if year is not None:
        records = [r for r in records if str(r.get("publication_year", "")).startswith(str(year))]

    kdc_dist: Counter[str] = Counter()
    mat_dist: Counter[str] = Counter()
    for r in records:
        kdc = r.get("kdc")
        if kdc:
            kdc_dist[str(kdc)[0]] += 1
        else:
            kdc_dist["?"] += 1
        # 자료유형은 ISBN/ISSN/electronic 패턴으로 추정
        if r.get("issn") or r.get("frequency"):
            mat_dist["SE"] += 1
        elif str(r.get("title", "")).startswith("[전자"):
            mat_dist["EB"] += 1
        else:
            mat_dist["BK"] += 1

    return len(records), dict(kdc_dist), dict(mat_dist)


def build_annual_stats(
    *,
    year: int,
    library_name: str,
    library_code: str = "",
    sourced_from_inventory: bool = True,
    overrides: dict[str, Any] | None = None,
) -> AnnualStats:
    """연간 통계 객체 생성. sourced_from_inventory=True면 자관 DB 자동 집계."""
    s = AnnualStats(year=year, library_name=library_name, library_code=library_code)
    if sourced_from_inventory:
        total, kdc_dist, mat_dist = compute_holdings_from_index(year=None)
        s.holdings_total = total
        s.holdings_by_kdc = kdc_dist
        s.holdings_by_material = mat_dist
        # 수서: 해당 연도 입력 자료 = 신규 구입 추정 (실제는 사서 보정)
        year_total, _, _ = compute_holdings_from_index(year=year)
        s.acquisition_purchase = year_total
    if overrides:
        for k, v in overrides.items():
            if hasattr(s, k):
                setattr(s, k, v)
    return s


def write_kolisnet_xlsx(
    stats: AnnualStats,
    output_path: str | Path,
) -> Path:
    """KOLIS-NET 제출용 연간 통계 엑셀.

    실제 제출 양식은 KOLIS-NET 시스템 내부 입력 화면이지만,
    사서가 본 엑셀을 보고 그대로 입력 가능한 형식.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:
        raise RuntimeError("openpyxl 미설치 — `pip install openpyxl`") from e

    wb = Workbook()

    # 시트 1: 요약
    ws = wb.active
    ws.title = "요약"
    header = Font(bold=True, size=12)
    label = Font(bold=True)
    fill = PatternFill("solid", fgColor="DDEEFF")

    row = 1
    cell = ws.cell(row=row, column=1, value=f"{stats.year}년 연간 통계")
    cell.font = header
    row += 1
    ws.cell(row=row, column=1, value="도서관명").font = label
    ws.cell(row=row, column=2, value=stats.library_name)
    row += 1
    ws.cell(row=row, column=1, value="KOLIS-NET 코드").font = label
    ws.cell(row=row, column=2, value=stats.library_code or "(미등록)")
    row += 2

    overview_rows = [
        ("총 장서", stats.holdings_total),
        ("연간 신규 수서 (구입)", stats.acquisition_purchase),
        ("기증 권수", stats.acquisition_donation),
        ("납본 권수", stats.acquisition_deposit),
        ("수서 비용 (원)", stats.acquisition_cost_krw),
        ("연간 대출", stats.loans_total),
        ("이용자 방문", stats.visits_total),
        ("등록 회원", stats.members_total),
        ("개관일", stats.operating_days),
        ("휴관일", stats.closed_days),
        ("정규직 사서", stats.librarians),
    ]
    for k, v in overview_rows:
        ws.cell(row=row, column=1, value=k).font = label
        ws.cell(row=row, column=2, value=v)
        row += 1

    # 시트 2: KDC 분포
    ws2 = wb.create_sheet("KDC 분포")
    ws2.append(["KDC 1자리", "권수"])
    for c in ws2[1]:
        c.font = label
        c.fill = fill
    for k in sorted(stats.holdings_by_kdc):
        ws2.append([k, stats.holdings_by_kdc[k]])

    # 시트 3: 자료유형
    ws3 = wb.create_sheet("자료유형")
    ws3.append(["코드", "명칭", "권수"])
    for c in ws3[1]:
        c.font = label
        c.fill = fill
    for code, name in MATERIAL_TYPES:
        ws3.append([code, name, stats.holdings_by_material.get(code, 0)])

    # 시트 4: 사서 보정 (빈 양식)
    ws4 = wb.create_sheet("보정 입력")
    ws4["A1"] = "KOLIS-NET 시스템에 직접 입력해야 하는 항목 (사서 보정)"
    ws4["A1"].font = label
    ws4["A2"] = "항목"
    ws4["B2"] = "값"
    ws4["C2"] = "비고"
    for c in ws4[2]:
        c.font = label
        c.fill = fill
    manual_rows = [
        ("연간 도서관 행사 횟수", "", "독서교실·작가초청·전시 등"),
        ("어린이실 운영 여부", "", "Y/N"),
        ("장애인 서비스 권수", "", "점자·녹음·대형활자 합산"),
        ("상호대차 신청 (책바다)", "", ""),
        ("상호대차 제공 (책바다)", "", ""),
        ("이용자 만족도 (5점)", "", "운영평가 지표"),
    ]
    for r in manual_rows:
        ws4.append(r)

    # 컬럼 폭
    from openpyxl.utils import get_column_letter

    for ws_ in (ws, ws2, ws3, ws4):
        for col_idx in range(1, 5):
            ws_.column_dimensions[get_column_letter(col_idx)].width = 22

    # 정렬
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2):
        for c in r:
            c.alignment = Alignment(horizontal="left")

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    logger.info("KOLIS-NET 연간통계 XLSX: %s", out)
    return out


def export_to_riss_for_school(
    stats: AnnualStats,
    output_path: str | Path,
) -> Path:
    """학교도서관진흥법 §6 — 한국교육학술정보원(KERIS) RISS 정보서비스시스템 연계 양식.

    학교도서관은 KOLIS-NET 외에 RISS 학교도서관 통합목록에도 통계 송신 의무.
    컬럼: 학교명·학교코드·총장서·신규수서·대출·이용자·전자책 등.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as e:
        raise RuntimeError("openpyxl 미설치") from e

    wb = Workbook()
    ws = wb.active
    ws.title = "RISS 학교도서관 통계"
    cols = [
        "학교명",
        "RISS 학교코드",
        "통계연도",
        "총 장서",
        "신규 수서",
        "전자책 (EB)",
        "기증 권수",
        "연간 대출",
        "회원 수",
        "정규직 사서 수",
        "비고",
    ]
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.append(
        [
            stats.library_name,
            stats.library_code,
            stats.year,
            stats.holdings_total,
            stats.acquisition_purchase,
            stats.holdings_by_material.get("EB", 0),
            stats.acquisition_donation,
            stats.loans_total,
            stats.members_total,
            stats.librarians,
            stats.notes,
        ]
    )

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    logger.info("RISS 학교도서관 통계 XLSX: %s", out)
    return out


__all__ = [
    "MATERIAL_TYPES",
    "AnnualStats",
    "build_annual_stats",
    "compute_holdings_from_index",
    "export_to_riss_for_school",
    "write_kolisnet_xlsx",
]
