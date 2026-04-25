"""상호대차 양식 어댑터.

PO 자료 흡수: 책나래 운영자 업무지침서(2023.3.) p3-15

각 시스템의 사서대리신청·일괄업로드 양식이 약간씩 다르므로 어댑터 패턴.
정확한 양식은 각 운영기관(NL Korea·KERIS)이 변경 가능. 본 모듈은 일반적
컬럼만 지원하며, 각 시스템에서 부분 import 후 사서가 수기 추가 컬럼 채울
수 있도록 빈 셀 보존.

매출 영향: 사서 1일 2회 점검 + 신청도서 KORMARC 매칭이 일과의 큰 부분.
우리 .mrc → 양식 자동 변환으로 1일 30분 절감.
"""

from __future__ import annotations

import csv
import io
import logging
from collections.abc import Iterable
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


# 책나래 사서대리신청 일괄업로드 컬럼 (운영자 업무지침서 §파일업로드)
CHAEKNARAE_COLUMNS = [
    "신청번호",
    "신청일자",
    "이용자명",
    "이용자ID",
    "도서명",
    "저자",
    "출판사",
    "발행년도",
    "ISBN",
    "청구기호",
    "등록번호",
    "보유여부",
    "비고",
]

# 책바다 신청 양식 (운영매뉴얼 2020)
CHAEKBADA_COLUMNS = [
    "신청번호",
    "신청일자",
    "신청자명",
    "신청자ID",
    "서명",  # 책바다는 "서명" 표기 (책나래 "도서명"과 다름)
    "저자",
    "출판사",
    "발행년도",
    "ISBN",
    "청구기호",
    "비고",
]

# RISS (KERIS) 상호대차 신청 양식
RISS_COLUMNS = [
    "신청번호",
    "신청일",
    "이용자",
    "소속기관",
    "자료명",
    "저자",
    "발행기관",
    "발행년도",
    "ISBN/ISSN",
    "권호",
    "페이지",
    "비고",
]


def _book_to_row(
    book: dict[str, Any],
    column_map: dict[str, list[str]],
) -> dict[str, str]:
    """book_data → 시스템별 row dict.

    column_map: {시스템 컬럼명: [book_data 키 후보 리스트]}
    """
    row: dict[str, str] = {}
    for col, candidates in column_map.items():
        value: str = ""
        for k in candidates:
            v = book.get(k)
            if v not in (None, ""):
                value = str(v)
                break
        row[col] = value
    return row


def _chaeknarae_map() -> dict[str, list[str]]:
    return {
        "신청번호": ["request_id", "id"],
        "신청일자": ["request_date", "date"],
        "이용자명": ["patron_name", "user_name"],
        "이용자ID": ["patron_id", "user_id"],
        "도서명": ["title"],
        "저자": ["author", "creator"],
        "출판사": ["publisher"],
        "발행년도": ["publication_year", "pub_year", "year"],
        "ISBN": ["isbn"],
        "청구기호": ["call_number", "call_no"],
        "등록번호": ["registration_number", "reg_no"],
        "보유여부": ["holding_status"],
        "비고": ["note", "remark"],
    }


def _chaekbada_map() -> dict[str, list[str]]:
    return {
        "신청번호": ["request_id", "id"],
        "신청일자": ["request_date", "date"],
        "신청자명": ["patron_name"],
        "신청자ID": ["patron_id"],
        "서명": ["title"],  # 책바다는 "서명"
        "저자": ["author"],
        "출판사": ["publisher"],
        "발행년도": ["publication_year", "pub_year"],
        "ISBN": ["isbn"],
        "청구기호": ["call_number"],
        "비고": ["note"],
    }


def _riss_map() -> dict[str, list[str]]:
    return {
        "신청번호": ["request_id"],
        "신청일": ["request_date"],
        "이용자": ["patron_name"],
        "소속기관": ["affiliation", "institution"],
        "자료명": ["title"],
        "저자": ["author"],
        "발행기관": ["publisher", "issuer"],
        "발행년도": ["publication_year"],
        "ISBN/ISSN": ["isbn", "issn"],
        "권호": ["volume", "issue"],
        "페이지": ["pages"],
        "비고": ["note"],
    }


def _convert(books: Iterable[dict[str, Any]], system: str) -> list[dict[str, str]]:
    if system == "chaeknarae":
        cmap = _chaeknarae_map()
        cols = CHAEKNARAE_COLUMNS
    elif system == "chaekbada":
        cmap = _chaekbada_map()
        cols = CHAEKBADA_COLUMNS
    elif system == "riss":
        cmap = _riss_map()
        cols = RISS_COLUMNS
    else:
        raise ValueError(f"알 수 없는 시스템: {system}")
    rows = [_book_to_row(b, cmap) for b in books]
    # 보장: 모든 컬럼 키 존재
    for r in rows:
        for c in cols:
            r.setdefault(c, "")
    return rows


def write_csv(
    books: Iterable[dict[str, Any]],
    output_path: str | Path,
    *,
    system: str,
) -> Path:
    """양식 CSV 저장 (UTF-8 BOM, Excel 호환)."""
    rows = _convert(books, system)
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    cols = {
        "chaeknarae": CHAEKNARAE_COLUMNS,
        "chaekbada": CHAEKBADA_COLUMNS,
        "riss": RISS_COLUMNS,
    }[system]

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=cols)
    writer.writeheader()
    writer.writerows(rows)
    out.write_bytes(buf.getvalue().encode("utf-8-sig"))
    logger.info("%s 양식 CSV 저장: %s (%d건)", system, out, len(rows))
    return out


def write_xlsx(
    books: Iterable[dict[str, Any]],
    output_path: str | Path,
    *,
    system: str,
) -> Path:
    """양식 Excel 저장 (헤더 굵게·자동 폭)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as e:
        raise RuntimeError("openpyxl 미설치 — `pip install openpyxl`") from e

    rows = _convert(books, system)
    cols = {
        "chaeknarae": CHAEKNARAE_COLUMNS,
        "chaekbada": CHAEKBADA_COLUMNS,
        "riss": RISS_COLUMNS,
    }[system]

    wb = Workbook()
    ws = wb.active
    label = {
        "chaeknarae": "책나래 사서대리신청",
        "chaekbada": "책바다 신청",
        "riss": "RISS 상호대차",
    }[system]
    ws.title = label[:31]  # 시트명 31자 제한

    ws.append(cols)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append([r.get(c, "") for c in cols])

    # 컬럼 폭
    for i, col in enumerate(cols, start=1):
        max_len = max(
            len(col),
            *(len(str(r.get(col, ""))) for r in rows),
        )
        from openpyxl.utils import get_column_letter

        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 40)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    logger.info("%s 양식 XLSX 저장: %s (%d건)", system, out, len(rows))
    return out


def from_inventory(
    isbns: Iterable[str],
) -> list[dict[str, Any]]:
    """자관 DB 인덱스에서 ISBN으로 도서 조회. 없으면 빈 dict.

    상호대차 신청 시 우리 자관 보유 여부 자동 표시용.
    """
    try:
        from kormarc_auto.inventory.library_db import iter_records
    except ImportError:
        return [{"isbn": x} for x in isbns]

    by_isbn: dict[str, dict[str, Any]] = {}
    for rec in iter_records():
        rn_isbn = str(rec.get("isbn") or "").replace("-", "")
        if rn_isbn:
            by_isbn[rn_isbn] = rec

    out: list[dict[str, Any]] = []
    for raw in isbns:
        clean = str(raw).replace("-", "").strip()
        rec = by_isbn.get(clean)
        if rec:
            out.append({**rec, "holding_status": "보유"})
        else:
            out.append({"isbn": clean, "holding_status": "미보유"})
    return out


__all__ = [
    "CHAEKBADA_COLUMNS",
    "CHAEKNARAE_COLUMNS",
    "RISS_COLUMNS",
    "from_inventory",
    "write_csv",
    "write_xlsx",
]
