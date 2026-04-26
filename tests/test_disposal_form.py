"""제적·폐기 결재서식 단위 테스트."""

from __future__ import annotations

import pytest

from kormarc_auto.output.disposal_form import (
    DISPOSAL_REASONS,
    DisposalEntry,
    render_disposal_form_pdf,
    write_disposal_xlsx,
)

SAMPLE = [
    DisposalEntry(
        registration_number="EM012600001",
        title="노후도서1",
        author="저자",
        reason_code="WORN",
        reason_detail="표지 분리",
        inspector="○○사서",
        cost_krw=12000,
    ),
    DisposalEntry(
        registration_number="EM012600002",
        title="복본도서",
        reason_code="DUPL",
        inspector="○○사서",
    ),
    DisposalEntry(
        registration_number="EM012600003",
        title="분실도서",
        reason_code="LOST",
        inspector="○○사서",
    ),
]


def test_disposal_reasons_complete():
    """대표 사유 코드 7종 모두 포함."""
    for code in ("WORN", "DUPL", "OBSO", "LOWUSE", "LOST", "DAMAGED", "REPLACE", "OTHER"):
        assert code in DISPOSAL_REASONS


def test_render_pdf(tmp_path):
    pytest.importorskip("reportlab")
    out = render_disposal_form_pdf(
        SAMPLE,
        library_name="테스트도서관",
        fiscal_period="2026 1분기",
        director="홍관장",
        output_path=tmp_path / "disp.pdf",
    )
    assert out.exists()
    assert out.read_bytes()[:4] == b"%PDF"


def test_render_pdf_empty(tmp_path):
    pytest.importorskip("reportlab")
    out = render_disposal_form_pdf(
        [],
        library_name="테스트도서관",
        fiscal_period="2026 1분기",
        output_path=tmp_path / "empty.pdf",
    )
    assert out.exists()


def test_write_xlsx(tmp_path):
    pytest.importorskip("openpyxl")
    out = write_disposal_xlsx(SAMPLE, tmp_path / "disp.xlsx")
    assert out.exists()
    from openpyxl import load_workbook

    wb = load_workbook(out)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert "등록번호" in header
    assert "사유코드" in header
    assert ws.max_row == 4  # 헤더 + 3건
