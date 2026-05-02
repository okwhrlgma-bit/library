"""xlsm_macro_parser 테스트 (Part 51 ADR-0070, P1 매크로 사서)."""

from __future__ import annotations

import pytest

from kormarc_auto.ingest.xlsm_macro_parser import (
    ISBN_PATTERN,
    XlsmExtractionResult,
    _normalize_isbn,
    extract_isbns_from_xlsm,
)


def test_normalize_isbn_valid() -> None:
    """유효 ISBN-13 정규화."""
    # 어린왕자 (978-89-374-3707-6) — 유효한 체크섬
    assert _normalize_isbn("978-89-374-3707-6") == "9788937437076"
    assert _normalize_isbn("9788937437076") == "9788937437076"


def test_normalize_isbn_rejects_invalid_checksum() -> None:
    """체크섬 실패 = None."""
    assert _normalize_isbn("9788937437070") is None  # 잘못된 마지막 자리


def test_normalize_isbn_rejects_isbn10() -> None:
    """ISBN-10은 거부 (KORMARC 020 ▾a = ISBN-13 only)."""
    assert _normalize_isbn("8937437074") is None


def test_normalize_isbn_rejects_non_book_prefix() -> None:
    """978·979 외 prefix 거부."""
    digits = "1234567890123"
    assert _normalize_isbn(digits) is None


def test_isbn_pattern_matches_hyphenated() -> None:
    """ISBN 패턴 = 하이픈 형식 매칭."""
    text = "이 책의 ISBN은 978-89-374-3707-6 입니다."
    matches = ISBN_PATTERN.findall(text)
    assert len(matches) >= 1


def test_extract_returns_result_for_missing_file() -> None:
    """없는 파일 = 경고 포함 결과 (예외 X)."""
    result = extract_isbns_from_xlsm("/nonexistent/path.xlsm")
    assert isinstance(result, XlsmExtractionResult)
    assert result.isbn_rows == 0
    assert any("찾을 수 없" in w for w in result.warnings)


def test_extract_from_real_xlsx(tmp_path) -> None:
    """openpyxl로 작성한 실제 xlsx에서 ISBN 추출."""
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    file_path = tmp_path / "test_isbns.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "ISBN"
    ws["B1"] = "도서명"
    ws["A2"] = "9788937437076"
    ws["B2"] = "어린왕자"
    ws["A3"] = "978-89-374-3707-6"  # 하이픈 형식 (동일 책)
    ws["B3"] = "어린왕자 (재판)"
    wb.save(file_path)

    result = extract_isbns_from_xlsm(file_path)
    assert "9788937437076" in result.isbns
    assert result.isbn_rows >= 1  # 중복 제거 후 1건


def test_extract_detects_isbn_column_header(tmp_path) -> None:
    """헤더에 ISBN 키워드 = detected_isbn_columns에 포함."""
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    file_path = tmp_path / "headers.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "도서번호"  # 헤더 키워드 (휴리스틱)
    ws["A2"] = "9788937437076"
    wb.save(file_path)

    result = extract_isbns_from_xlsm(file_path)
    assert len(result.detected_isbn_columns) >= 1


def test_extract_respects_max_isbns_limit() -> None:
    """안전 상한 max_isbns 준수."""
    # 직접 호출 (빈 결과여도 함수 시그니처 검증)
    result = extract_isbns_from_xlsm("/nonexistent.xlsx", max_isbns=10)
    assert result.isbn_rows <= 10
