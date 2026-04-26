"""연간 KOLIS-NET 통계 단위 테스트."""

from __future__ import annotations

from unittest.mock import patch

import pytest

from kormarc_auto.output.annual_statistics import (
    MATERIAL_TYPES,
    AnnualStats,
    build_annual_stats,
    compute_holdings_from_index,
    export_to_riss_for_school,
    write_kolisnet_xlsx,
)


def test_material_types_includes_book_serial_eb():
    codes = [c for c, _ in MATERIAL_TYPES]
    assert "BK" in codes
    assert "SE" in codes
    assert "EB" in codes


@patch("kormarc_auto.inventory.library_db.search_local")
def test_compute_holdings_aggregates(mock_search):
    mock_search.return_value = [
        {"kdc": "813.7", "publication_year": "2026"},
        {"kdc": "813.6", "publication_year": "2026"},
        {"kdc": "911.05", "publication_year": "2025"},
        {"issn": "1234-5678"},  # 연속간행물
    ]
    total, kdc_dist, mat_dist = compute_holdings_from_index()
    assert total == 4
    assert kdc_dist["8"] == 2
    assert kdc_dist["9"] == 1
    assert mat_dist["SE"] == 1


@patch("kormarc_auto.inventory.library_db.search_local")
def test_build_annual_stats_with_overrides(mock_search):
    mock_search.return_value = [{"kdc": "813.7", "publication_year": "2026"}]
    s = build_annual_stats(
        year=2026,
        library_name="테스트도서관",
        overrides={"loans_total": 100, "librarians": 2},
    )
    assert s.library_name == "테스트도서관"
    assert s.year == 2026
    assert s.loans_total == 100
    assert s.librarians == 2
    assert s.holdings_total >= 1


def test_write_kolisnet_xlsx(tmp_path):
    pytest.importorskip("openpyxl")
    s = AnnualStats(
        year=2026,
        library_name="○○도서관",
        library_code="LIB1234",
        holdings_total=5000,
        holdings_by_kdc={"8": 1500, "9": 800},
        holdings_by_material={"BK": 4500, "EB": 500},
        acquisition_purchase=300,
        loans_total=12000,
        members_total=800,
    )
    out = write_kolisnet_xlsx(s, tmp_path / "stats.xlsx")
    assert out.exists()
    from openpyxl import load_workbook

    wb = load_workbook(out)
    assert "요약" in wb.sheetnames
    assert "KDC 분포" in wb.sheetnames
    assert "자료유형" in wb.sheetnames
    assert "보정 입력" in wb.sheetnames


def test_export_riss_for_school(tmp_path):
    pytest.importorskip("openpyxl")
    s = AnnualStats(
        year=2026,
        library_name="○○고등학교도서관",
        library_code="SCH9999",
        holdings_total=3000,
        holdings_by_material={"BK": 2700, "EB": 300},
    )
    out = export_to_riss_for_school(s, tmp_path / "riss.xlsx")
    assert out.exists()
    from openpyxl import load_workbook

    wb = load_workbook(out)
    ws = wb.active
    assert ws["A2"].value == "○○고등학교도서관"
