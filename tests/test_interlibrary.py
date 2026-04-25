"""상호대차 양식 어댑터 단위 테스트."""

from __future__ import annotations

import csv
import io

import pytest

from kormarc_auto.interlibrary.exporters import (
    CHAEKNARAE_COLUMNS,
    RISS_COLUMNS,
    write_csv,
    write_xlsx,
)

SAMPLE_BOOKS = [
    {
        "title": "작별하지 않는다",
        "author": "한강",
        "publisher": "문학동네",
        "publication_year": "2021",
        "isbn": "9788954682152",
        "patron_name": "이용자A",
        "request_id": "R20260415-001",
        "request_date": "2026-04-15",
        "call_number": "813.7-한강",
        "registration_number": "EM012600123",
        "holding_status": "미보유",
        "note": "이용자 우선 신청",
    },
    {
        "title": "삼대",
        "author": "염상섭",
        "publisher": "한국문학사",
        "publication_year": "1932",
        "isbn": "9788912345678",
    },
]


def test_write_csv_chaeknarae(tmp_path):
    out = write_csv(SAMPLE_BOOKS, tmp_path / "n.csv", system="chaeknarae")
    assert out.exists()
    text = out.read_bytes().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    assert len(rows) == 2
    assert reader.fieldnames == CHAEKNARAE_COLUMNS
    assert rows[0]["도서명"] == "작별하지 않는다"
    assert rows[0]["청구기호"] == "813.7-한강"
    assert rows[0]["등록번호"] == "EM012600123"
    assert rows[0]["보유여부"] == "미보유"
    assert rows[1]["등록번호"] == ""  # 누락은 빈 셀


def test_write_csv_chaekbada_uses_seomyeong(tmp_path):
    """책바다는 '서명' 컬럼 (도서명 X)."""
    out = write_csv(SAMPLE_BOOKS, tmp_path / "b.csv", system="chaekbada")
    text = out.read_bytes().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    assert "서명" in reader.fieldnames
    assert "도서명" not in reader.fieldnames
    rows = list(reader)
    assert rows[0]["서명"] == "작별하지 않는다"


def test_write_csv_riss_columns(tmp_path):
    out = write_csv(SAMPLE_BOOKS, tmp_path / "r.csv", system="riss")
    text = out.read_bytes().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    assert reader.fieldnames == RISS_COLUMNS
    rows = list(reader)
    assert rows[0]["자료명"] == "작별하지 않는다"


def test_write_csv_unknown_system(tmp_path):
    with pytest.raises(ValueError):
        write_csv(SAMPLE_BOOKS, tmp_path / "x.csv", system="unknown")


def test_write_xlsx_chaeknarae(tmp_path):
    pytest.importorskip("openpyxl")
    out = write_xlsx(SAMPLE_BOOKS, tmp_path / "n.xlsx", system="chaeknarae")
    assert out.exists()
    from openpyxl import load_workbook

    wb = load_workbook(out)
    ws = wb.active
    assert ws.title.startswith("책나래")
    header = [c.value for c in ws[1]]
    assert header == CHAEKNARAE_COLUMNS
